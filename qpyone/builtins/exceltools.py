#!/usr/bin/env python
from typing import Any

from pathlib import Path

import pyexcel

from openpyxl import load_workbook
from pydantic import BaseModel


__all__ = ["read_as_objects", "write_objects_to_file"]


def read_as_objects(
    excel_path: str | Path,
    model: type[BaseModel] = BaseModel,
    sheet_index: int = 0,
    sheet_name: str = None,
    ignore_validate_errors: bool = False,
) -> list[Any]:
    wb = load_workbook(excel_path)
    if sheet_name is not None:
        sheet = wb.get_sheet_by_name(sheet_name)
    else:
        sheet = wb.worksheets[sheet_index]
    rows = sheet.iter_rows()
    headers = {th.value: i for i, th in enumerate(next(rows))}
    objects = []
    for row_index, raw_row in enumerate(rows, start=2):
        try:
            item = {key: v.value for key, v in zip(headers, raw_row) if key is not None}
            objects.append(model.parse_obj(item))
        except BaseException as e:
            if ignore_validate_errors:
                continue
            raise e
    return objects


def write_objects_to_file(objects: list[BaseModel | dict | dict], excel_path: str):
    """
    write objects to excel or csv
    """
    if isinstance(objects[0], dict) or isinstance(objects[0], dict):
        pyexcel.get_sheet(records=objects).save_as(excel_path)
    elif isinstance(objects[0], BaseModel):
        pyexcel.get_sheet(
            records=[item.dict(by_alias=True) for item in objects]
        ).save_as(excel_path)
    else:
        raise NotImplementedError("Not Support Model type")
