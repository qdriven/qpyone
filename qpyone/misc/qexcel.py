#!/usr/bin/env python
from typing import Any

from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel

from .excel import XlsxModel


"""
handle excel
"""


def read_as_objects(
    excel_path: str | Path,
    model: type[BaseModel] = BaseModel,
    sheet_index: int = 0,
    sheet_name: str = None,
    ignore_validate_errors: bool = False,
) -> list[Any]:
    wb = load_workbook(excel_path)
    if sheet_name is not None:
        sheet = wb.get_sheet_by_name(sheet_name)
    else:
        sheet = wb.worksheets[sheet_index]
    rows = sheet.iter_rows()
    headers = {th.value: i for i, th in enumerate(next(rows))}
    objects = []
    for row_index, raw_row in enumerate(rows, start=2):
        try:
            item = {key: v.value for key, v in zip(headers, raw_row) if key is not None}
            objects.append(model.parse_obj(item))
        except BaseException as e:
            if ignore_validate_errors:
                continue
            raise e
    return objects


def to_excel(excel_path: str | Path, data: XlsxModel):
    data.to_file(excel_path)
